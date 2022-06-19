# -*- coding:UTF-8 -*-

# datetime:2022/6/15 13:12
# software: PyCharm
"""
//
//                       .::::.
//                     .::::::::.
//                    :::::::::::
//                 ..:::::::::::'
//              '::::::::::::'
//                .::::::::::
//           '::::::::::::::..
//                ..::::::::::::.
//              ``::::::::::::::::
//               ::::``:::::::::'        .:::.
//              ::::'   ':::::'       .::::::::.
//            .::::'      ::::     .:::::::'::::.
//           .:::'       :::::  .:::::::::' ':::::.
//          .::'        :::::.:::::::::'      ':::::.
//         .::'         ::::::::::::::'         ``::::.
//     ...:::           ::::::::::::'              ``::.
//    ```` ':.          ':::::::::'                  ::::..
//                       '.:::::'                    ':'````..
"""
"""
文件说明：
    
"""
from openpyxl import load_workbook
from django.shortcuts import render, redirect, HttpResponse
from app01 import models


def depart_list(request):
    """ 部门列表 """

    # 去数据库获取所有的部门信息
    # [obj, obj]
    queryset = models.Department.objects.all()

    return render(request, 'depart_list.html', {'queryset': queryset})


def depart_add(request):
    """ 添加部门 """

    if request.method == 'GET':
        return render(request, 'depart_add.html')

    # 获取用户POST提交过来的数据
    title = request.POST.get('title')

    # 保存到数据库
    models.Department.objects.create(title=title)

    # 重定向回到部门列表
    return redirect('/depart/list/')


def depart_delete(request):
    """ 删除部门 """
    # 获取id
    # http://127.0.0.1:8000/depart/delete/?nid=1
    nid = request.GET.get('nid')
    # 删除
    models.Department.objects.filter(id=nid).delete()
    # 跳转回部门列表
    return redirect('/depart/list')


# http://127.0.0.1:8000/depart/1/edit/
def depart_edit(request, nid):
    """ 修改部门 """
    if request.method == "GET":
        # 根据nid，获取它的数据 【obj】
        row_object = models.Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {"row_object": row_object})

    # 获取用户提交的标题
    title = request.POST.get('title')
    # 根据ID找到数据库中的数据并进行更新
    models.Department.objects.filter(id=nid).update(title=title)
    # 重定向回到部门列表
    return redirect('/depart/list/')


def depart_multi(request):
    """" 批量上传（Excel文件） """
    file_object = request.FILES.get("exc")
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect('/depart/list/')